import requests
import pandas as pd
from matplotlib import pyplot as plt
import numpy as np
from openpyxl import *
import os
from openpyxl import load_workbook


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

API =  "RGAPI-546ff69b-0cba-44ac-b79d-be1472f09bc1"
Region = "EUW1"
summonerName = "Klaus147 "

SummonerInfo = requests.get("https://" + Region + ".api.riotgames.com/lol/summoner/v4/summoners/by-name/" + summonerName + "?api_key=" + API)
print(SummonerInfo.json())
print(SummonerInfo)
SummonerInfo = SummonerInfo.json()
id =  SummonerInfo['id'] 
RankedMatches = requests.get("https://" + Region + ".api.riotgames.com/lol/league/v4/entries/by-summoner/" + id + "?api_key=" + API)
Ranked = RankedMatches.json()
print(Ranked)
print("https://" + Region + ".api.riotgames.com/lol/league/v4/entries/by-summoner/" + id + "?api_key=" + API)
#print(RankedMatches)
rankee = Ranked[0]
#print(rankee)
masteryScore = requests.get("https://" + Region + ".api.riotgames.com/lol/champion-mastery/v4/champion-masteries/by-summoner/" + id + "?api_key=" + API)
#print(masteryScore.json())

masteryScore = masteryScore.json()
sortedScore = sorted(masteryScore, key=lambda k: k['championPoints'], reverse=True)

MatchIDs = requests.get("https://europe.api.riotgames.com/lol/match/v5/matches/by-puuid/"+ SummonerInfo['puuid'] +  "/ids?start=0&count=5&api_key=" + API)
count = 5
MatchIDs = MatchIDs.json()



data = {
                'GameDuration':[], 
                'champion': [],
                'kills':[] ,
                'deaths': [],
                'assists': [],
                'goldEarned':[],
                'physicalDamageDealtToChampions':[],
                'physicalDamageTaken':[],
                'cs':[],
                'dragonKills':[],
                'baronKills':[],
                'win': []
        }


for matchID in MatchIDs:
    MatchData = requests.get("https://europe.api.riotgames.com/lol/match/v5/matches/"+ matchID +"?api_key=RGAPI-546ff69b-0cba-44ac-b79d-be1472f09bc1")
    #print("https://europe.api.riotgames.com/lol/match/v5/matches/"+ matchID +"?api_key=RGAPI-49316441-6393-4a47-b249-d06ceb6fc9fe")
    MatchData = MatchData.json()
    # A list of all the participants puuids
    participants = MatchData['metadata']['participants']
    # Now, find where in the data our players puuid is found
    player_index = participants.index(SummonerInfo['puuid'])
    #print(MatchData['info']['participants'][player_index])
    player_data = MatchData['info']['participants'][player_index]
    gameMins = MatchData['info']['gameDuration']
    champion = player_data['championName']
    k = player_data['kills']
    d = player_data['deaths']
    a = player_data['assists']
    win = player_data['win']
    GoldPerMin = player_data['goldEarned']
    physicalDamageDealtToChampions = player_data['physicalDamageDealtToChampions']
    physicalDamageTaken = player_data['physicalDamageTaken']
    cs = player_data['totalMinionsKilled']
    dragonKills = player_data['dragonKills']
    baronKills = player_data['baronKills']
    data['champion'].append(champion)
    data['kills'].append(k)
    data['deaths'].append(d)
    data['assists'].append(a)
    data['win'].append(win)    
    data['GameDuration'].append(gameMins)
    data['goldEarned'].append(GoldPerMin)
    data['physicalDamageDealtToChampions'].append(physicalDamageDealtToChampions)
    data['physicalDamageTaken'].append(physicalDamageTaken)
    data['cs'].append(cs)
    data['dragonKills'].append(dragonKills)    
    data['baronKills'].append(baronKills)



MeanData = {
        'avgGoldPerMin': [],
        'creepScore': [],
        'totalDamageDonePerMin': [],
        'totalDamageTakenPerMin': []
}

ListS = []
ie = 0
for matchID in MatchIDs:
        data2 = {
                'currentGold': [],
                'minionsKilled': [],
                'totalDamageDoneToChampions': [],
                'totalDamageTaken': []
        }

        MatchData = requests.get("https://europe.api.riotgames.com/lol/match/v5/matches/"+ matchID +"/timeline?api_key=RGAPI-546ff69b-0cba-44ac-b79d-be1472f09bc1")
        print("https://europe.api.riotgames.com/lol/match/v5/matches/"+ matchID +"/timeline?api_key=RGAPI-546ff69b-0cba-44ac-b79d-be1472f09bc1")
        MatchData = MatchData.json()
        participants = MatchData['metadata']['participants']
        # Now, find where in the data our players puuid is found
        player_index = participants.index(SummonerInfo['puuid'])
        player_index = str(player_index +1 )
        matchData = MatchData['info']['frames']

        i = 0
        matchL = data['GameDuration'][ie] /60 + 1
        ie =  ie + 1 
        while i < matchL:
                player_data = MatchData['info']['frames'][i]['participantFrames'][player_index]
                currentGold = player_data['totalGold']
                minionsKilled = player_data['minionsKilled']
                
                totalDamageDoneToChampions = player_data['damageStats']['totalDamageDoneToChampions']
                totalDamageTaken = player_data['damageStats']['totalDamageTaken']
                
                data2['currentGold'].append(currentGold)
                data2['minionsKilled'].append(minionsKilled)
                data2['totalDamageDoneToChampions'].append(totalDamageDoneToChampions)
                data2['totalDamageTaken'].append(totalDamageTaken)
                i= i + 1   
        
        ListS.append(data2)
        print(ListS)
        print("==============================")

i = 0
counter = 0
mean = 0


plt.clf()
df = pd.DataFrame(data)    
df['win'] = df['win'].astype(int) 
print(df)



with pd.ExcelWriter('data.xlsx', engine='openpyxl', mode='a') as writer: 
     df.to_excel(writer) 




#print(df)

